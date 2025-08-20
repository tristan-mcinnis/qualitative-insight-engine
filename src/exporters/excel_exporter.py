from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from src.config.config_loader import AppConfig

class ExcelFormatter:
    def __init__(self, config: AppConfig):
        self.config = config
    
    def save_dataframe(self, df: pd.DataFrame, filename: Path, 
                      sheet_name: str = "Report") -> None:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            self._format_worksheet(writer.sheets[sheet_name])
    
    def save_comprehensive_report(self, sheets_data: dict[str, pd.DataFrame], 
                                filename: Path) -> None:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, 
                           index=(sheet_name == 'Cross-Tabulation'))
                self._format_worksheet(writer.sheets[sheet_name], sheet_name)
    
    def _format_worksheet(self, worksheet, sheet_name: str = None) -> None:
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color=self.config.excel.header_color,
            end_color=self.config.excel.header_color,
            fill_type="solid"
        )
        
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Add borders for cross-tabulation
        if sheet_name == 'Cross-Tabulation':
            self._add_borders(worksheet)
        
        # Auto-fit columns
        self._auto_fit_columns(worksheet, sheet_name)
        
        # Wrap text
        self._wrap_text(worksheet)
    
    def _add_borders(self, worksheet) -> None:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.rows:
            for cell in row:
                cell.border = thin_border
                if cell.column == 1:
                    cell.font = Font(bold=True)
    
    def _auto_fit_columns(self, worksheet, sheet_name: str = None) -> None:
        for column in worksheet.columns:
            max_length = 0
            column_letter = None
            
            try:
                column_letter = column[0].column_letter
            except (AttributeError, IndexError):
                continue
            
            sample_cells = list(column)[:100]
            
            for cell in sample_cells:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            if (sheet_name == 'Master Analysis' and 
                column_letter in ['G', 'H', 'I']):
                worksheet.column_dimensions[column_letter].width = (
                    self.config.excel.verbatim_column_width
                )
            else:
                adjusted_width = min(
                    max_length + 2, 
                    self.config.excel.max_column_width
                )
                worksheet.column_dimensions[column_letter].width = (
                    max(adjusted_width, 10)
                )
    
    def _wrap_text(self, worksheet) -> None:
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')